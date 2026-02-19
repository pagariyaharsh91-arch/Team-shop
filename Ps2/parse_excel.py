#!/usr/bin/env python3
"""
parse_excel.py
Parses STOCK.xlsx and outputs products as JSON for further processing.
Requires: openpyxl or pandas
"""

import json
import sys

try:
    # Try openpyxl first (lightweight)
    from openpyxl import load_workbook
    
    wb = load_workbook('STOCK.xlsx')
    ws = wb.active
    
    products = []
    headers = {}
    
    # Read headers from first row
    for col_idx, cell in enumerate(ws[1], 1):
        headers[col_idx] = cell.value
    
    # Read data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):  # Skip empty rows
            continue
        
        # Map cells to headers
        name = row[0] if len(row) > 0 else ''
        category = row[1] if len(row) > 1 else 'Uncategorized'
        price = row[2] if len(row) > 2 else 0
        quantity = row[3] if len(row) > 3 else ''
        
        # Clean values
        name = str(name).strip() if name else ''
        category = str(category).strip() if category else 'Uncategorized'
        
        try:
            price = float(str(price).replace('₹', '').strip()) if price else 0
        except:
            price = 0
        
        try:
            quantity = int(str(quantity).strip()) if quantity and str(quantity).strip() else ''
        except:
            quantity = ''
        
        if name:
            products.append({
                'name': name,
                'category': category,
                'price': price,
                'quantity': quantity
            })
    
    # Group by category
    grouped = {}
    for p in products:
        cat = p['category']
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(p)
    
    output = {
        'products': products,
        'grouped': grouped,
        'total_products': len(products),
        'total_categories': len(grouped)
    }
    
    print(json.dumps(output, indent=2, ensure_ascii=False))

except ImportError:
    # Fallback: try pandas
    try:
        import pandas as pd
        df = pd.read_excel('STOCK.xlsx')
        
        # Normalize column names
        df.columns = [str(c).strip() for c in df.columns]
        
        products = []
        for _, row in df.iterrows():
            name = str(row.iloc[0]).strip() if len(row) > 0 else ''
            category = str(row.iloc[1]).strip() if len(row) > 1 else 'Uncategorized'
            price = row.iloc[2] if len(row) > 2 else 0
            quantity = row.iloc[3] if len(row) > 3 else ''
            
            try:
                price = float(str(price).replace('₹', '').strip()) if price else 0
            except:
                price = 0
            
            if name:
                products.append({
                    'name': name,
                    'category': category,
                    'price': price,
                    'quantity': quantity if quantity != '' else ''
                })
        
        # Group by category
        grouped = {}
        for p in products:
            cat = p['category']
            if cat not in grouped:
                grouped[cat] = []
            grouped[cat].append(p)
        
        output = {
            'products': products,
            'grouped': grouped,
            'total_products': len(products),
            'total_categories': len(grouped)
        }
        
        print(json.dumps(output, indent=2, ensure_ascii=False))
    
    except Exception as e:
        print(json.dumps({'error': str(e)}), file=sys.stderr)
        sys.exit(1)

except Exception as e:
    print(json.dumps({'error': str(e)}), file=sys.stderr)
    sys.exit(1)
